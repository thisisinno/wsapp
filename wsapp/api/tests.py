import io
import json
import base64
from pathlib import Path
from datetime import timedelta
from unittest.mock import Mock, patch

import requests
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.conf import settings
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .models import (
    Campaign,
    CampaignRecipient,
    ImportedRecipient,
    MessageAttempt,
    MessagingPreference,
    UploadedDataset,
    UploadedMedia,
)
from .services.campaigns import SendIntervalError, check_recipient, parse_send_interval, queue_campaign_entries
from .services.imports import disambiguate_headers, parse_tabular
from .services.phones import excel_value_to_text, normalize_tanzania_phone
from .services.templates import TemplateError, render_message, validate_template
from .services.wasender import (
    MalformedResponseError,
    ProviderResult,
    RateLimitError,
    UnauthorizedError,
    ValidationError,
    WasenderClient,
    WasenderError,
    normalize_message_status,
)
from .services.file_types import MediaValidationError, resolve_and_validate_media

User = get_user_model()


def make_dataset(owner, content=b"phone,name\n0712345678,Ada\n", name="x.csv"):
    return UploadedDataset.objects.create(
        owner=owner,
        original_file=SimpleUploadedFile(name, content),
        original_filename=name,
        file_type=name.rsplit(".", 1)[-1],
        size=len(content),
        checksum="a" * 64,
        processing_status="ready",
        detected_columns=[{"key": "phone", "label": "Phone", "index": 1}],
        selected_phone_column="phone",
    )


class PhoneImportTemplateTests(TestCase):
    def test_normalization_import_and_templates(self):
        for value in [
            "+255712345678",
            "255712345678",
            "0712345678",
            "712345678",
            "0712 345-678",
            "0712345678.0",
        ]:
            self.assertEqual(normalize_tanzania_phone(value).normalized, "+255712345678")
        self.assertEqual(excel_value_to_text(712345678.0), "712345678")
        self.assertEqual(
            [item["key"] for item in disambiguate_headers(["Name", "", "Name"])],
            ["name", "column_2", "name_2"],
        )
        columns, rows = parse_tabular(
            io.BytesIO(b"Name,Phone\nAda,0712345678\n"), "x.csv"
        )
        self.assertEqual((len(columns), rows[0][1]["phone"]), (2, "0712345678"))
        self.assertEqual(validate_template("Hi {name}", {"name"}), ["name"])
        with self.assertRaises(TemplateError):
            validate_template("{missing}", {"name"})
        self.assertEqual(
            render_message("Hi {name} {{ok}}", {"name": "Ada"})[0], "Hi Ada {ok}"
        )


@override_settings(
    WASENDER_API_KEY="test-key-never-used",
    WASENDER_SEND_INTERVAL_SECONDS=0,
)
class CampaignSequentialTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("owner", password="pw")
        self.other = User.objects.create_user("other", password="pw")
        self.dataset = make_dataset(self.user)
        for index in range(4):
            ImportedRecipient.objects.create(
                dataset=self.dataset,
                owner=self.user,
                original_row_number=index + 2,
                row_data={"phone": f"071234567{index}", "name": f"Person {index}"},
                phone_source_column="phone",
                phone_original=f"071234567{index}",
                phone_normalized=f"+25571234567{index}",
                phone_validation_status="valid",
            )
        self.campaign = Campaign.objects.create(
            owner=self.user,
            dataset=self.dataset,
            name="Four",
            body_snapshot="Hi {name}",
            selected_phone_column="phone",
            opt_in_confirmed=True,
            allow_unknown=True,
            send_interval_seconds=0,
            status=Campaign.Status.READY,
        )
        self.client.force_login(self.user)
        self.start_url = reverse("campaign_action", args=[self.campaign.id, "start"])
        self.next_url = reverse("campaign_send_next", args=[self.campaign.id])

    def start(self):
        return self.client.post(
            self.start_url, data="{}", content_type="application/json"
        )

    def next(self, token):
        return self.client.post(
            self.next_url,
            data=json.dumps({"run_token": token}),
            content_type="application/json",
        )

    @patch("api.services.campaigns.WasenderClient.check_number")
    @patch("api.services.campaigns.WasenderClient.send_message")
    def test_start_and_four_one_item_requests(self, send, check):
        check.return_value = ProviderResult({"data": {"exists": True}}, 200)
        send.return_value = ProviderResult({"success": True, "data": {"msgId": "m", "status": "in_progress"}}, 200)
        first = self.start()
        self.assertEqual(first.status_code, 200)
        data = first.json()["data"]
        self.assertEqual((data["progress_text"], data["total"]), ("0/4", 4))
        self.assertEqual(self.campaign.campaign_recipients.count(), 4)
        self.assertEqual(send.call_count, 0)
        token = data["run_token"]
        for number in range(1, 5):
            response = self.next(token)
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.json()["data"]["progress_text"], f"{number}/4")
            self.assertEqual(send.call_count, number)
            CampaignRecipient.objects.filter(campaign=self.campaign).exclude(
                attempt_started_at__isnull=True
            ).update(attempt_started_at=timezone.now() - timedelta(seconds=6))
        self.assertTrue(response.json()["data"]["finished"])

    @patch("api.services.campaigns.WasenderClient.check_number")
    @patch("api.services.campaigns.WasenderClient.send_message")
    def test_validation_failure_is_processed_once(self, send, check):
        check.return_value = ProviderResult({"data": {"exists": True}}, 200)
        send.side_effect = ValidationError("invalid destination", 422, {"safe": True})
        token = self.start().json()["data"]["run_token"]
        data = self.next(token).json()["data"]
        self.assertEqual((data["processed"], data["failed"]), (1, 1))
        self.assertEqual(send.call_count, 1)
        self.assertEqual(MessageAttempt.objects.count(), 1)

    @patch("api.services.campaigns.WasenderClient.check_number")
    @patch("api.services.campaigns.WasenderClient.send_message")
    def test_campaign_interval_prevents_immediate_second_provider_call(self, send, check):
        self.campaign.send_interval_seconds = 60
        self.campaign.save()
        check.return_value = ProviderResult({"data": {"exists": True}}, 200)
        send.return_value = ProviderResult({"success": True, "data": {"msgId": "m"}}, 200)
        token = self.start().json()["data"]["run_token"]
        self.next(token)
        second = self.next(token).json()["data"]
        self.assertFalse(second["sent_now"])
        self.assertGreater(second["wait_seconds"], 0)
        self.assertEqual(send.call_count, 1)

    @patch("api.services.campaigns.WasenderClient.send_message")
    def test_provider_protection_requeues_without_permanent_failure(self, send):
        send.side_effect = RateLimitError(
            "You have account protection enabled. You can only send 1 message every 5 seconds.",
            429,
            {"retry_after": 7},
            7,
        )
        token = self.start().json()["data"]["run_token"]
        response = self.next(token)
        self.assertEqual(response.status_code, 200)
        data = response.json()["data"]
        self.assertTrue(data["rate_limited"])
        self.assertGreaterEqual(data["wait_seconds"], 7)
        entry = self.campaign.campaign_recipients.order_by("sequence_number").first()
        entry.refresh_from_db()
        self.assertEqual(entry.state, "queued")
        self.assertIsNone(entry.failed_at)
        self.assertEqual(entry.attempts.last().error_category, "rate_limited")

    @patch("api.services.campaigns.WasenderClient.check_number")
    @patch("api.services.campaigns.WasenderClient.send_message")
    def test_idempotent_start_resume_token_and_no_duplicate_accepted(self, send, check):
        check.return_value = ProviderResult({"data": {"exists": True}}, 200)
        send.return_value = ProviderResult({"success": True, "data": {"msgId": "m"}}, 200)
        first = self.start().json()["data"]
        second = self.start().json()["data"]
        self.assertNotEqual(first["run_token"], second["run_token"])
        self.assertEqual(self.campaign.campaign_recipients.count(), 4)
        self.next(second["run_token"])
        CampaignRecipient.objects.filter(campaign=self.campaign).exclude(
            attempt_started_at__isnull=True
        ).update(attempt_started_at=timezone.now() - timedelta(seconds=6))
        self.client.post(
            reverse("campaign_action", args=[self.campaign.id, "pause"]),
            data="{}",
            content_type="application/json",
        )
        resumed = self.client.post(
            reverse("campaign_action", args=[self.campaign.id, "resume"]),
            data="{}",
            content_type="application/json",
        ).json()["data"]
        self.next(resumed["run_token"])
        self.assertEqual(send.call_count, 2)
        self.assertEqual(
            self.campaign.campaign_recipients.filter(state="accepted").count(), 2
        )

    @patch("api.services.campaigns.WasenderClient.send_message")
    def test_mismatched_token_pause_cancel_and_non_whatsapp(self, send):
        token = self.start().json()["data"]["run_token"]
        self.assertEqual(self.next("wrong").status_code, 409)
        self.assertEqual(send.call_count, 0)
        pause = self.client.post(
            reverse("campaign_action", args=[self.campaign.id, "pause"]),
            data="{}",
            content_type="application/json",
        )
        self.assertEqual(pause.status_code, 200)
        self.assertEqual(self.next(token).status_code, 409)
        self.client.post(
            reverse("campaign_action", args=[self.campaign.id, "cancel"]),
            data="{}",
            content_type="application/json",
        )
        self.assertEqual(send.call_count, 0)

    @patch("api.services.campaigns.WasenderClient.send_message")
    def test_send_does_not_require_number_precheck(self, send):
        send.return_value = ProviderResult({"success": True, "data": {"msgId": "m"}}, 200)
        token = self.start().json()["data"]["run_token"]
        data = self.next(token).json()["data"]
        self.assertEqual(data["accepted"], 1)
        self.assertEqual(send.call_count, 1)
        self.assertEqual(
            self.campaign.campaign_recipients.order_by("sequence_number")
            .first()
            .state,
            "accepted",
        )

    @patch("api.services.campaigns.WasenderClient.check_number")
    def test_fresh_database_check_reused(self, check):
        recipient = self.dataset.recipients.first()
        check.return_value = ProviderResult({"data": {"exists": True}}, 200)
        self.assertEqual(check_recipient(recipient), "exists")
        recipient.refresh_from_db()
        self.assertEqual(check_recipient(recipient), "exists")
        self.assertEqual(check.call_count, 1)

    def test_progress_masks_and_owner_scope(self):
        self.start()
        response = self.client.get(
            reverse("campaign_progress", args=[self.campaign.id])
        )
        self.assertEqual(response["Cache-Control"], "no-store")
        text = response.content.decode()
        self.assertNotIn("+255712345670", text)
        self.assertNotIn("test-key-never-used", text)
        self.client.force_login(self.other)
        for name, method in [
            ("campaign_progress", self.client.get),
            ("campaign_send_next", self.client.post),
            ("campaign_action", self.client.post),
        ]:
            args = (
                [self.campaign.id, "start"]
                if name == "campaign_action"
                else [self.campaign.id]
            )
            self.assertEqual(method(reverse(name, args=args)).status_code, 404)

    def test_cancel_only_changes_queued_and_resend_only_failed(self):
        token = self.start().json()["data"]["run_token"]
        entries = list(self.campaign.campaign_recipients.order_by("sequence_number"))
        entries[0].state = "accepted"
        entries[0].save()
        entries[1].state = "failed"
        entries[1].skip_reason = "bad"
        entries[1].save()
        response = self.client.post(
            reverse("resend_failed", args=[self.campaign.id]),
            data="{}",
            content_type="application/json",
        )
        self.assertEqual(response.json()["data"]["requeued"], 1)
        entries[0].refresh_from_db()
        self.assertEqual(entries[0].state, "accepted")
        self.client.post(
            reverse("campaign_action", args=[self.campaign.id, "cancel"]),
            data="{}",
            content_type="application/json",
        )
        entries[0].refresh_from_db()
        self.assertEqual(entries[0].state, "accepted")

    def test_export_and_cross_user_access(self):
        queue_campaign_entries(self.campaign)
        response = self.client.get(reverse("export_campaign", args=[self.campaign.id]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            load_workbook(io.BytesIO(response.content)).active["A1"].value, "row"
        )

    @patch("api.services.campaigns.WasenderClient.send_message")
    def test_unexpected_exception_marks_failed_and_never_leaks_key(self, send):
        send.side_effect = RuntimeError(
            f"unexpected detail {settings.WASENDER_API_KEY}"
        )
        token = self.start().json()["data"]["run_token"]
        with self.assertLogs("api.services.campaigns", level="ERROR") as captured:
            data = self.next(token).json()["data"]
        entry = self.campaign.campaign_recipients.order_by("sequence_number").first()
        entry.refresh_from_db()
        self.assertEqual((data["failed"], entry.state), (1, "failed"))
        self.assertIsNotNone(entry.attempt_finished_at)
        self.assertNotIn(settings.WASENDER_API_KEY, json.dumps(data))
        # logger.exception includes the original exception, so errors must not
        # interpolate provider payloads or credentials into exception messages.
        self.assertIn("Unexpected provider processing failure", captured.output[0])

    @patch("api.services.campaigns.WasenderClient.send_message")
    def test_busy_claim_does_not_make_provider_call(self, send):
        token = self.start().json()["data"]["run_token"]
        first = self.campaign.campaign_recipients.order_by("sequence_number").first()
        first.state = CampaignRecipient.State.PROCESSING
        first.attempt_started_at = timezone.now()
        first.save()
        data = self.next(token).json()["data"]
        self.assertTrue(data["busy"])
        self.assertEqual(send.call_count, 0)


class SynchronousUploadMediaTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("files", password="pw")
        self.client.force_login(self.user)

    def test_csv_upload_completes_synchronously(self):
        response = self.client.post(
            reverse("upload_create"),
            {"file": SimpleUploadedFile("people.csv", b"phone,name\n0712345678,Ada\n")},
        )
        self.assertEqual(response.status_code, 200)
        dataset = UploadedDataset.objects.get(owner=self.user)
        self.assertEqual((dataset.processing_status, dataset.row_count), ("ready", 1))

    @patch("api.services.media.WasenderClient.upload_media")
    def test_media_upload_completes_synchronously(self, upload):
        upload.return_value = ProviderResult({"publicUrl": "https://cdn.invalid/a.jpg"}, 200)
        response = self.client.post(
            reverse("media_create"),
            {"file": SimpleUploadedFile("a.jpg", b"image", content_type="image/jpeg")},
        )
        self.assertEqual(response.status_code, 200)
        media = UploadedMedia.objects.get(owner=self.user)
        self.assertEqual(media.upload_status, "ready")
        self.assertTrue(media.original_file)


class ExcelMediaRegressionTests(TestCase):
    """All upload calls are mocked: these tests never contact the provider."""
    def setUp(self):
        self.user = User.objects.create_user("excel-files", password="pw")
        self.client.force_login(self.user)

    @staticmethod
    def xlsx_bytes():
        workbook = Workbook()
        workbook.active["A1"] = "students"
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def post_media(self, name, content, content_type=""):
        return self.client.post(reverse("media_create"), {
            "file": SimpleUploadedFile(name, content, content_type=content_type),
        })

    @patch("api.services.media.WasenderClient.upload_media")
    def test_xlsx_accepts_generic_and_zip_browser_mime_and_stores_canonical_mime(self, upload):
        upload.return_value = ProviderResult({"publicUrl": "https://cdn.invalid/a.xlsx"}, 200)
        for mime in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/octet-stream", "application/zip", ""):
            response = self.post_media("students.xlsx", self.xlsx_bytes(), mime)
            self.assertEqual(response.status_code, 200, response.content)
        media = UploadedMedia.objects.filter(owner=self.user).latest("created_at")
        self.assertEqual(media.mime_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertEqual(media.original_filename, "students.xlsx")

    @patch("api.services.media.WasenderClient.upload_media")
    def test_csv_pdf_xls_and_safe_basename(self, upload):
        upload.return_value = ProviderResult({"publicUrl": "https://cdn.invalid/document"}, 200)
        self.assertEqual(self.post_media("../../people.csv", b"name,phone\nAda,0712\n", "text/csv").status_code, 200)
        self.assertEqual(self.post_media("report.pdf", b"%PDF-1.4\nbody", "application/pdf").status_code, 200)
        self.assertEqual(self.post_media("legacy.xls", bytes.fromhex("D0CF11E0A1B11AE1") + b"legacy", "application/vnd.ms-excel").status_code, 200)
        media = UploadedMedia.objects.filter(owner=self.user, original_filename="people.csv").get()
        self.assertEqual(media.original_filename, "people.csv")

    def test_corrupt_or_arbitrary_zip_xlsx_gets_field_error(self):
        for content in (b"not a workbook", b"PK\x03\x04not-a-real-zip"):
            response = self.post_media("bad.xlsx", content, "application/zip")
            self.assertEqual(response.status_code, 400)
            self.assertIn("file", response.json()["errors"])
            self.assertIn("code", response.json()["data"])

    @patch("api.services.media.WasenderClient.upload_media")
    def test_provider_rejection_returns_safe_field_error(self, upload):
        upload.side_effect = ValidationError("provider detail that is not returned", 422, {})
        response = self.post_media("students.xlsx", self.xlsx_bytes(), "application/zip")
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["data"]["code"], "provider_media_validation")
        self.assertEqual(response.json()["errors"]["file"], ["The provider rejected this media type."])

    @override_settings(MEDIA_DOCUMENT_MAX_BYTES=10, MEDIA_IMAGE_MAX_BYTES=3)
    def test_category_size_policy_uses_document_limit(self):
        with self.assertRaises(MediaValidationError):
            resolve_and_validate_media(SimpleUploadedFile("large.jpg", b"1234", content_type="image/jpeg"))
        self.assertEqual(
            resolve_and_validate_media(SimpleUploadedFile("small.csv", b"1234", content_type="text/csv")).category,
            "document",
        )
        with self.assertRaises(MediaValidationError):
            resolve_and_validate_media(SimpleUploadedFile("large.csv", b"12345678901", content_type="text/csv"))


class ProviderClientTests(TestCase):
    def response(self, status, payload=None, text=""):
        response = Mock(status_code=status, ok=status < 400, headers={}, text=text)
        if payload is None:
            response.json.side_effect = ValueError()
        else:
            response.json.return_value = payload
        return response

    def client_for(self, response=None, exc=None):
        session = Mock()
        session.request.side_effect = exc
        if not exc:
            session.request.return_value = response
        return WasenderClient(
            api_key="example", base_url="https://example.invalid", session=session
        )

    def test_success_errors_timeout_and_non_json(self):
        result = self.client_for(
            self.response(200, {"success": True, "data": {"msgId": 1}})
        ).send_message("+255712345678", "Hi")
        self.assertEqual(result.http_status, 200)
        for status, exception in [
            (401, UnauthorizedError),
            (422, ValidationError),
            (429, RateLimitError),
            (500, WasenderError),
        ]:
            with self.assertRaises(exception):
                self.client_for(self.response(status, {"message": "safe"})).check_number(
                    "+255712345678"
                )
        from .services.wasender import TimeoutError

        with self.assertRaises(TimeoutError):
            self.client_for(exc=requests.Timeout()).check_number("+255712345678")
        with self.assertRaises(MalformedResponseError):
            self.client_for(self.response(200, None, "not json")).check_number(
                "+255712345678"
            )

    def test_ooxml_raw_upload_uses_canonical_mime_and_one_base64_fallback(self):
        content = ExcelMediaRegressionTests.xlsx_bytes()
        session = Mock()
        session.request.side_effect = [
            self.response(422, {"message": "invalid media type"}),
            self.response(200, {"success": True, "publicUrl": "https://cdn.invalid/workbook"}),
        ]
        client = WasenderClient(api_key="example", base_url="https://example.invalid", session=session)
        result = client.upload_media(io.BytesIO(content), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx")
        self.assertEqual(result.http_status, 200)
        self.assertEqual(session.request.call_count, 2)
        self.assertEqual(session.request.call_args_list[0].kwargs["headers"]["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        fallback = session.request.call_args_list[1].kwargs["json"]
        self.assertEqual(fallback["mimetype"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertEqual(base64.b64decode(fallback["base64"]), content)

    def test_document_filename_only_goes_to_document_payload(self):
        session = Mock()
        session.request.return_value = self.response(200, {"success": True, "data": {}})
        client = WasenderClient(api_key="example", base_url="https://example.invalid", session=session)
        client.send_message("+255712345678", "Report", "documentUrl", "https://cdn.invalid/a", "students.xlsx")
        self.assertEqual(session.request.call_args.kwargs["json"]["fileName"], "students.xlsx")
        client.send_message("+255712345678", "Photo", "imageUrl", "https://cdn.invalid/i", "photo.jpg")
        self.assertNotIn("fileName", session.request.call_args.kwargs["json"])


class CsrfAndFrontendTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("csrf", password="pw")
        self.dataset = make_dataset(self.user)
        self.client = Client(enforce_csrf_checks=True)
        self.client.force_login(self.user)

    def payload(self):
        return json.dumps({
            "name": "CSRF",
            "body": "Hello",
            "opt_in_confirmed": True,
        })

    def test_port_9000_is_trusted_and_valid_ajax_create_succeeds(self):
        self.assertIn("https://localhost:9000", settings.CSRF_TRUSTED_ORIGINS)
        page = self.client.get(reverse("campaign_new", args=[self.dataset.id]))
        self.assertEqual(page.status_code, 200)
        token = self.client.cookies["csrftoken"].value
        response = self.client.post(
            reverse("campaign_create", args=[self.dataset.id]),
            self.payload(),
            content_type="application/json",
            HTTP_ORIGIN="https://localhost:9000",
            HTTP_X_CSRFTOKEN=token,
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 200)

    def test_missing_csrf_is_rejected_without_creating_campaign(self):
        response = self.client.post(
            reverse("campaign_create", args=[self.dataset.id]),
            self.payload(),
            content_type="application/json",
            HTTP_ORIGIN="https://localhost:9000",
        )
        self.assertEqual(response.status_code, 403)
        self.assertFalse(Campaign.objects.filter(owner=self.user).exists())

    def test_javascript_handles_html_403_and_uses_same_origin_credentials(self):
        source = (Path(settings.BASE_DIR) / "static/api/app.js").read_text()
        self.assertIn('credentials: "same-origin"', source)
        self.assertIn('contentType.includes("application/json")', source)
        self.assertIn("Security check failed. Refresh the page and retry.", source)
        self.assertNotIn("const payload = await response.json();", source)

    def test_webhook_url_is_removed(self):
        self.assertEqual(self.client.post("/webhooks/wasender/").status_code, 404)


class SendIntervalTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("interval", password="pw")
        self.dataset = make_dataset(self.user)
        self.client.force_login(self.user)
        self.preference = MessagingPreference.objects.create(
            owner=self.user, default_send_interval_seconds=30
        )

    def create(self, interval_marker=None):
        payload = {"name": "Interval", "body": "Hello", "opt_in_confirmed": True}
        if interval_marker is not None:
            payload["send_interval_seconds"] = interval_marker
        return self.client.post(
            reverse("campaign_create", args=[self.dataset.id]),
            data=json.dumps(payload), content_type="application/json",
        )

    def test_strict_parser_accepts_zero_and_one(self):
        self.assertEqual(parse_send_interval("0"), 0)
        self.assertEqual(parse_send_interval(0), 0)
        self.assertEqual(parse_send_interval("1"), 1)

    def test_strict_parser_rejects_invalid_values(self):
        for value in ("-1", "1.5", "1e2", "words", "3601"):
            with self.assertRaises(SendIntervalError):
                parse_send_interval(value)

    @patch("api.services.wasender.WasenderClient.send_message")
    @patch("api.services.wasender.WasenderClient.check_number")
    def test_campaign_interval_defaults_and_zero_are_preserved(self, check, send):
        for value, expected in ((None, 30), ("", 30), (0, 0), ("0", 0)):
            response = self.create(value)
            self.assertEqual(response.status_code, 200)
            campaign = Campaign.objects.get(pk=response.json()["data"]["id"])
            self.assertEqual(campaign.send_interval_seconds, expected)
            self.assertEqual(campaign.send_config_snapshot["interval_seconds"], expected)
        self.assertFalse(check.called)
        self.assertFalse(send.called)

    def test_campaign_invalid_intervals_are_field_errors(self):
        for value in ("-1", "1.5", "3601"):
            response = self.create(value)
            self.assertEqual(response.status_code, 400)
            self.assertIn("send_interval_seconds", response.json()["errors"])

    def test_settings_zero_and_blank_preserve_values(self):
        url = reverse("messaging_settings_save")
        response = self.client.post(url, json.dumps({"default_send_interval_seconds": 0}), content_type="application/json")
        self.assertEqual(response.status_code, 200)
        self.preference.refresh_from_db()
        self.assertEqual(self.preference.default_send_interval_seconds, 0)
        self.preference.default_send_interval_seconds = 60
        self.preference.save()
        response = self.client.post(url, json.dumps({"default_send_interval_seconds": ""}), content_type="application/json")
        self.assertEqual(response.status_code, 200)
        self.preference.refresh_from_db()
        self.assertEqual(self.preference.default_send_interval_seconds, 60)

    def test_interval_templates_and_versioned_assets(self):
        campaign = self.client.get(reverse("campaign_new", args=[self.dataset.id])).content.decode()
        settings_page = self.client.get(reverse("messaging_settings")).content.decode()
        base = (Path(settings.BASE_DIR) / "templates/base.html").read_text()
        javascript = (Path(settings.BASE_DIR) / "static/api/app.js").read_text()
        self.assertNotIn("required", campaign.split('id="sendInterval"', 1)[1].split(">", 1)[0])
        self.assertNotIn("required", settings_page.split('id="defaultInterval"', 1)[1].split(">", 1)[0])
        self.assertIn('min="0"', campaign)
        self.assertIn('min="0"', settings_page)
        self.assertNotIn("Custom seconds (5–3600)", settings_page)
        self.assertNotIn("provider enforces a minimum of 5 seconds", settings_page)
        self.assertIn("api/app.js' %}?v={{ STATIC_ASSET_VERSION }}", base)
        self.assertIn("api/app.css' %}?v={{ STATIC_ASSET_VERSION }}", base)
        self.assertIn("parseIntervalInput", javascript)
        self.assertIn("^(0|[1-9]\\d*)$", javascript)
        self.assertNotIn("interval.value ||", javascript)
        self.assertNotIn("Number(interval.value) ||", javascript)
        self.assertIn('id="sendIntervalFeedback" class="invalid-feedback" hidden', campaign)


@override_settings(
    WASENDER_API_KEY="test-key-never-used",
    WASENDER_SEND_INTERVAL_SECONDS=0,
)
class MessageLogTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("logs", password="pw")
        self.other = User.objects.create_user("intruder", password="pw")
        self.dataset = make_dataset(self.user)
        self.recipient = ImportedRecipient.objects.create(
            dataset=self.dataset,
            owner=self.user,
            original_row_number=2,
            row_data={"phone": "0712345678"},
            phone_source_column="phone",
            phone_original="0712345678",
            phone_normalized="+255712345678",
            phone_validation_status="valid",
        )
        self.campaign = Campaign.objects.create(
            owner=self.user,
            dataset=self.dataset,
            name="Owner campaign",
            body_snapshot="Hello",
            selected_phone_column="phone",
            send_interval_seconds=0,
        )
        self.entry = CampaignRecipient.objects.create(
            campaign=self.campaign,
            imported_recipient=self.recipient,
            normalized_phone="+255712345678",
            rendered_message="Hello Ada",
            state="failed",
            provider_message_id="provider-1",
            skip_reason="Previous failure",
        )
        other_dataset = make_dataset(self.other, name="other.csv")
        other_recipient = ImportedRecipient.objects.create(
            dataset=other_dataset,
            owner=self.other,
            original_row_number=2,
            phone_normalized="+255713000000",
        )
        other_campaign = Campaign.objects.create(
            owner=self.other,
            dataset=other_dataset,
            name="Private campaign",
            body_snapshot="Secret message",
            selected_phone_column="phone",
        )
        CampaignRecipient.objects.create(
            campaign=other_campaign,
            imported_recipient=other_recipient,
            normalized_phone="+255713000000",
            rendered_message="Secret message",
            state="sent",
        )
        self.client.force_login(self.user)

    def post(self, name, payload=None):
        return self.client.post(
            reverse(name, args=[self.entry.id]),
            data=json.dumps(payload or {}),
            content_type="application/json",
        )

    def test_status_normalization_is_conservative(self):
        expected = {
            1: "pending",
            2: "sent",
            3: "delivered",
            4: "read",
            5: "played",
            99: "unknown",
        }
        for code, state in expected.items():
            self.assertEqual(normalize_message_status(code), state)
        self.assertEqual(normalize_message_status(None, "in_progress"), "pending")
        self.assertEqual(normalize_message_status(None, "invented"), "unknown")

    def test_numeric_data_status_and_text_aliases_normalize(self):
        self.assertEqual(normalize_message_status(3), "delivered")
        self.assertEqual(normalize_message_status(None, "server_acknowledged"), "sent")
        self.assertEqual(normalize_message_status(None, "failure"), "failed")

    def test_page_owner_scope_filters_and_serials(self):
        page = self.client.get(reverse("message_logs"))
        text = page.content.decode()
        self.assertContains(page, "Owner campaign")
        self.assertNotIn("Private campaign", text)
        self.assertNotIn("Secret message", text)
        self.assertContains(page, "+255712345678")
        self.assertContains(page, 'class="message-serial">1')
        self.assertEqual(
            self.client.get(reverse("message_logs"), {"phone": "999"}).context["page_obj"].paginator.count,
            0,
        )
        self.assertEqual(
            self.client.get(reverse("message_logs"), {"status": "failed"}).context["page_obj"].paginator.count,
            1,
        )
        self.assertEqual(
            self.client.get(reverse("message_logs"), {"campaign": self.campaign.id}).context["page_obj"].paginator.count,
            1,
        )

    def test_detail_has_safe_attempt_diagnostics_and_owner_scope(self):
        MessageAttempt.objects.create(
            campaign_recipient=self.entry,
            attempt_number=1,
            http_status=422,
            error_category="validation",
            error_message="Safe diagnostic",
            provider_response={"authorization": settings.WASENDER_API_KEY, "message": settings.WASENDER_API_KEY},
            duration_ms=17,
        )
        response = self.client.get(reverse("message_detail", args=[self.entry.id]))
        payload = response.json()["data"]
        self.assertEqual(payload["attempts"][0]["error_message"], "Safe diagnostic")
        self.assertNotIn(settings.WASENDER_API_KEY, response.content.decode())
        self.client.force_login(self.other)
        self.assertEqual(
            self.client.get(reverse("message_detail", args=[self.entry.id])).status_code,
            404,
        )

    @patch("api.services.message_logs.WasenderClient.message_info")
    def test_refresh_advances_timestamps_and_never_downgrades(self, info):
        self.entry.state = "read"
        self.entry.save()
        info.return_value = ProviderResult(
            {"success": True, "data": {"ack": 2, "status": "sent"}}, 200
        )
        response = self.post("message_refresh_status")
        self.assertEqual(response.status_code, 200)
        self.entry.refresh_from_db()
        self.assertEqual(self.entry.state, "read")
        self.assertIsNotNone(self.entry.sent_at)
        info.return_value = ProviderResult(
            {"success": True, "data": {"ack": 3, "status": "delivered"}}, 200
        )
        self.entry.state = "sent"
        self.entry.delivered_at = None
        self.entry.save()
        self.post("message_refresh_status")
        self.entry.refresh_from_db()
        self.campaign.refresh_from_db()
        self.assertEqual(self.entry.state, "delivered")
        self.assertIsNotNone(self.entry.delivered_at)
        self.assertEqual(self.campaign.delivered_count, 1)

    @patch("api.services.message_logs.WasenderClient.edit_message")
    def test_provider_edit_success_and_failure(self, edit):
        edit.return_value = ProviderResult({"success": True, "data": {}}, 200)
        self.assertEqual(self.post("message_update", {"text": "Corrected"}).status_code, 200)
        self.entry.refresh_from_db()
        self.assertEqual(self.entry.rendered_message, "Corrected")
        edit.side_effect = WasenderError("Safe rejection", 400, {"safe": True})
        self.assertEqual(self.post("message_update", {"text": "Must not save"}).status_code, 400)
        self.entry.refresh_from_db()
        self.assertEqual(self.entry.rendered_message, "Corrected")

    def test_failed_without_provider_id_can_be_corrected_locally(self):
        self.entry.provider_message_id = ""
        self.entry.save()
        response = self.post(
            "message_update",
            {"phone": "0711111111", "text": "Try this", "update_imported_recipient": True},
        )
        self.assertEqual(response.status_code, 200)
        self.entry.refresh_from_db()
        self.recipient.refresh_from_db()
        self.assertEqual(self.entry.normalized_phone, "+255711111111")
        self.assertEqual(self.recipient.phone_normalized, "+255711111111")

    @patch("api.services.message_logs.WasenderClient.delete_message")
    def test_delete_preserves_state_and_blocks_repeat(self, delete):
        delete.return_value = ProviderResult({"success": True}, 200)
        self.assertEqual(self.post("message_delete").status_code, 200)
        self.entry.refresh_from_db()
        self.assertEqual(self.entry.state, "failed")
        self.assertIsNotNone(self.entry.provider_deleted_at)
        self.assertEqual(self.post("message_delete").status_code, 409)
        self.assertEqual(delete.call_count, 1)

    @patch("api.services.message_logs.WasenderClient.send_message")
    @patch("api.services.message_logs.WasenderClient.resend_message")
    def test_resend_selects_operation_and_records_success(self, resend, send):
        resend.return_value = ProviderResult(
            {"success": True, "data": {"msgId": "provider-2", "status": "pending"}}, 200
        )
        response = self.post(
            "message_resend",
            {"phone": self.entry.normalized_phone, "text": self.entry.rendered_message},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual((resend.call_count, send.call_count), (1, 0))
        self.entry.refresh_from_db()
        self.assertEqual((self.entry.state, self.entry.retry_count), ("pending", 1))
        self.assertEqual(self.entry.attempts.count(), 1)
        self.entry.state = "failed"
        self.entry.failed_at = timezone.now()
        self.entry.attempt_started_at = timezone.now() - timedelta(seconds=6)
        self.entry.save()
        send.return_value = ProviderResult(
            {"success": True, "data": {"msgId": "fresh", "status": "in_progress"}}, 200
        )
        response = self.post(
            "message_resend",
            {"phone": "0711111111", "text": "Changed"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(send.call_count, 1)

    @patch("api.services.message_logs.WasenderClient.send_message")
    def test_failed_resend_preserves_corrections_and_diagnostics(self, send):
        self.entry.provider_message_id = ""
        self.entry.save()
        send.side_effect = WasenderError("Invalid corrected destination", 422, {"safe": True})
        response = self.post(
            "message_resend",
            {"phone": "0711111111", "text": "Corrected body"},
        )
        self.assertEqual(response.status_code, 422)
        self.entry.refresh_from_db()
        attempt = self.entry.attempts.get()
        self.assertEqual(self.entry.rendered_message, "Corrected body")
        self.assertEqual(attempt.error_message, "Invalid corrected destination")

    @patch("api.services.message_logs.WasenderClient.resend_message")
    def test_campaign_interval_makes_no_provider_call(self, resend):
        self.campaign.send_interval_seconds = 60
        self.campaign.save()
        self.entry.attempt_started_at = timezone.now()
        self.entry.save()
        MessageAttempt.objects.create(
            campaign_recipient=self.entry,
            attempt_number=1,
        )
        response = self.post(
            "message_resend",
            {"phone": self.entry.normalized_phone, "text": self.entry.rendered_message},
        )
        self.assertEqual(response.status_code, 429)
        self.assertGreater(response.json()["data"]["wait_seconds"], 0)
        resend.assert_not_called()

    def test_unknown_old_action_is_404(self):
        self.assertEqual(
            self.client.post(f"/messages/{self.entry.id}/nonsense/").status_code,
            404,
        )

    def test_message_mutation_requires_csrf(self):
        csrf_client = Client(enforce_csrf_checks=True)
        csrf_client.force_login(self.user)
        response = csrf_client.post(
            reverse("message_update", args=[self.entry.id]),
            data=json.dumps({"text": "No token"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 403)
        self.assertEqual(Campaign.objects.filter(owner=self.user).count(), 1)

    @patch("api.views.WasenderClient.message_info")
    def test_auto_sync_advances_and_returns_compact_patch(self, info):
        self.entry.state = "sent"
        self.entry.save()
        info.return_value = ProviderResult({"success": True, "data": {"status": 3}}, 200)
        response = self.client.post(reverse("message_auto_sync_statuses"), data=json.dumps({"ids": [str(self.entry.id)]}), content_type="application/json")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Cache-Control"], "no-store")
        result = response.json()["data"]["results"][0]
        self.assertTrue(result["changed"])
        self.assertEqual(result["patch"]["state"], "delivered")
        self.assertNotIn("row", result)
        self.assertIn("campaign_counts", response.json()["data"])
        self.entry.refresh_from_db()
        self.assertIsNotNone(self.entry.delivered_at)

    @patch("api.views.WasenderClient.message_info")
    def test_auto_sync_throttles_and_provider_error_preserves_state(self, info):
        self.entry.state = "sent"
        self.entry.next_status_check_at = timezone.now() + timedelta(seconds=30)
        self.entry.save()
        response = self.client.post(reverse("message_auto_sync_statuses"), data=json.dumps({"ids": [str(self.entry.id)]}), content_type="application/json")
        self.assertEqual(response.json()["data"]["checked"], 0)
        info.assert_not_called()
        self.entry.next_status_check_at = None
        self.entry.save()
        info.side_effect = WasenderError("temporary outage", 503, {})
        response = self.client.post(reverse("message_auto_sync_statuses"), data=json.dumps({"ids": [str(self.entry.id)]}), content_type="application/json")
        self.assertEqual(response.status_code, 200)
        self.entry.refresh_from_db()
        self.assertEqual(self.entry.state, "sent")
        self.assertTrue(self.entry.status_sync_error)
        self.assertIsNotNone(self.entry.next_status_check_at)

    def test_auto_sync_requires_login_csrf_and_does_not_expose_other_rows(self):
        url = reverse("message_auto_sync_statuses")
        self.client.logout()
        self.assertEqual(self.client.post(url, data="{}", content_type="application/json").status_code, 302)
        csrf_client = Client(enforce_csrf_checks=True); csrf_client.force_login(self.user)
        self.assertEqual(csrf_client.post(url, data="{}", content_type="application/json").status_code, 403)
        self.client.force_login(self.user)
        private = CampaignRecipient.objects.filter(campaign__owner=self.other).first()
        response = self.client.post(url, data=json.dumps({"ids": [str(private.id)]}), content_type="application/json")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["data"]["results"], [])


class ProviderAcceptanceTests(ProviderClientTests):
    def test_2xx_success_false_is_failure(self):
        with self.assertRaises(WasenderError):
            self.client_for(
                self.response(200, {"success": False, "message": "not accepted"})
            ).send_message("+255712345678", "Hi")

    def test_authorization_is_session_only_and_payload_is_minimal(self):
        session = Mock()
        session.headers = {}
        session.request.return_value = self.response(
            200, {"success": True, "data": {"msgId": 1}}
        )
        WasenderClient(
            api_key="server-secret",
            base_url="https://example.invalid",
            session=session,
        ).send_message("+255712345678", "Hi")
        self.assertEqual(session.headers["Authorization"], "Bearer server-secret")
        kwargs = session.request.call_args.kwargs
        self.assertNotIn("headers", kwargs)
        self.assertEqual(kwargs["json"], {"to": "+255712345678", "text": "Hi"})
