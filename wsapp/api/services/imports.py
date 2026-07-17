import csv
import hashlib
import io
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from .phones import excel_value_to_text


class ImportError(ValueError):
    pass


def checksum(upload):
    digest = hashlib.sha256()
    for chunk in upload.chunks():
        digest.update(chunk)
    upload.seek(0)
    return digest.hexdigest()


def safe_value(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return excel_value_to_text(value)
    if isinstance(value, str):
        return value
    return str(value)


def disambiguate_headers(values):
    used, result = {}, []
    for index, value in enumerate(values, 1):
        label = str(value).strip() if value is not None else ""
        label = label or f"Column {index}"
        base = "".join(c.lower() if c.isalnum() else "_" for c in label).strip("_") or f"column_{index}"
        key = base
        used[base] = used.get(base, 0) + 1
        if used[base] > 1:
            key = f"{base}_{used[base]}"
        result.append({"key": key, "label": label, "index": index})
    return result


def workbook_sheets(fileobj, suffix):
    if suffix == ".csv":
        return ["CSV"]
    if suffix == ".xls":
        try:
            import xlrd
            book = xlrd.open_workbook(file_contents=fileobj.read())
            fileobj.seek(0)
            return book.sheet_names()
        except Exception as exc:
            raise ImportError(f"Could not read legacy Excel file: {exc}") from exc
    try:
        book = load_workbook(fileobj, read_only=True, data_only=True)
        return book.sheetnames
    except Exception as exc:
        raise ImportError(f"Could not read workbook: {exc}") from exc


def parse_tabular(fileobj, filename, sheet_name="", header_row=1):
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = io.TextIOWrapper(fileobj, encoding="utf-8-sig", errors="replace", newline="")
        values = list(csv.reader(text))
        text.detach()
    elif suffix == ".xlsx":
        book = load_workbook(fileobj, read_only=True, data_only=True)
        sheet = book[sheet_name] if sheet_name else book[book.sheetnames[0]]
        values = list(sheet.iter_rows(values_only=True))
    elif suffix == ".xls":
        import xlrd
        book = xlrd.open_workbook(file_contents=fileobj.read())
        sheet = book.sheet_by_name(sheet_name) if sheet_name else book.sheet_by_index(0)
        values = [sheet.row_values(i) for i in range(sheet.nrows)]
    else:
        raise ImportError("Only .xlsx, .xls and .csv files are supported.")
    if not values or len(values) < header_row:
        raise ImportError("The file does not contain the selected header row.")
    columns = disambiguate_headers(values[header_row - 1])
    rows = []
    for number, cells in enumerate(values[header_row:], header_row + 1):
        mapped = {column["key"]: safe_value(cells[i] if i < len(cells) else "") for i, column in enumerate(columns)}
        if any(str(v).strip() for v in mapped.values()):
            rows.append((number, mapped))
    return columns, rows
